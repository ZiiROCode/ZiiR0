import os
import re
import logging
import datetime
import pytz
import json  
import asyncio 
import asyncpg 
from typing import Optional 
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from telegram import Update
from telegram.ext import Application, CommandHandler, MessageHandler, filters, ContextTypes

logging.basicConfig(
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
    level=logging.INFO
)
logger = logging.getLogger(__name__)

TELEGRAM_TOKEN = os.environ.get("TELEGRAM_TOKEN", "YOUR_TELEGRAM_BOT_TOKEN")
ADMIN_USERNAME = os.environ.get("ADMIN_USERNAME", "excelmerge")
CAMBODIA_TZ = pytz.timezone('Asia/Phnom_Penh')

DATABASE_URL = os.environ.get("DATABASE_URL")
if not DATABASE_URL:
    logger.error("FATAL: DATABASE_URL environment variable is not set.")

DB_POOL: Optional[asyncpg.Pool] = None
db_lock = asyncio.Lock()

user_data = {}
user_breaks = {}

async def get_db_pool() -> asyncpg.Pool:
    """Gets the shared database connection pool."""
    global DB_POOL
    if DB_POOL is None or DB_POOL.is_closing():
        if not DATABASE_URL:
            raise ValueError("DATABASE_URL is not set, cannot create pool.")
        try:
            DB_POOL = await asyncpg.create_pool(
                dsn=DATABASE_URL,
                max_inactive_connection_lifetime=60,
                min_size=1,
                max_size=5 
            )
            if DB_POOL is None:
                raise ConnectionError("Database pool initialization failed.")
            logger.info("Database connection pool established.")
        except Exception as e:
            logger.error(f"Could not create database connection pool: {e}")
            raise
    return DB_POOL

async def close_db_pool():
    """Closes the database connection pool."""
    global DB_POOL
    if DB_POOL and not DB_POOL.is_closing():
        logger.info("Closing database connection pool.")
        await DB_POOL.close()
        DB_POOL = None

def _norm_owner_name(s: str) -> str:
    """Normalizes an owner name (copied from other bot)."""
    s = (s or "").strip()
    if s.startswith("@"): s = s[1:]
    return s.lower()

async def _set_owner_status_in_db(owner_name: str, is_stopped: bool):
    """Finds an owner in the 'owners' JSON blob and sets their disabled status."""
    if not owner_name:
        logger.warning("Attempted to set owner status for an empty username.")
        return
    
    normalized_name = _norm_owner_name(owner_name)
    if not normalized_name:
        return

    async with db_lock:
        try:
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                async with conn.transaction():
                    result = await conn.fetchval("SELECT data FROM kv_storage WHERE key = 'owners' FOR UPDATE")
                    
                    if not result:
                        logger.warning(f"Could not find 'owners' key in kv_storage.")
                        return

                    owner_data = json.loads(result)
                    found_owner = False
                    for owner_group in owner_data:
                        if _norm_owner_name(owner_group.get("owner", "")) == normalized_name:
                            owner_group["disabled"] = is_stopped
                            if not is_stopped:
                                owner_group.pop("disabled_until", None)
                            found_owner = True
                            break
                    
                    if found_owner:
                        await conn.execute("UPDATE kv_storage SET data = $1, updated_at = NOW() WHERE key = 'owners'", json.dumps(owner_data))
                        await conn.execute("NOTIFY owners_changed;")
                        logger.info(f"Set owner '{normalized_name}' status to {'STOPPED' if is_stopped else 'OPENED'}.")
                    else:
                        logger.info(f"User '{normalized_name}' checked in/out, but is not a configured owner in kv_storage.")

        except Exception as e:
            logger.error(f"Failed to set owner status for '{normalized_name}': {e}", exc_info=True)

async def _stop_all_owners_in_db():
    """Sets all owners in the 'owners' JSON blob to disabled: true."""
    logger.info("Running daily job to stop all owners...")
    async with db_lock:
        try:
            pool = await get_db_pool()
            async with pool.acquire() as conn:
                async with conn.transaction():
                    result = await conn.fetchval("SELECT data FROM kv_storage WHERE key = 'owners' FOR UPDATE")
                    if not result:
                        logger.warning(f"Could not find 'owners' key in kv_storage for daily stop.")
                        return
                    
                    owner_data = json.loads(result)
                    changes_made = False
                    for owner_group in owner_data:
                        if not owner_group.get("disabled", False):
                            owner_group["disabled"] = True
                            changes_made = True
                    
                    if changes_made:
                        await conn.execute("UPDATE kv_storage SET data = $1, updated_at = NOW() WHERE key = 'owners'", json.dumps(owner_data))
                        await conn.execute("NOTIFY owners_changed;")
                        logger.info("Stopped all owners for end-of-day.")
                    else:
                        logger.info("All owners were already stopped. No changes made.")

        except Exception as e:
            logger.error(f"Failed to stop all owners: {e}", exc_info=True)

def get_now():
    """Gets the current time in Cambodia timezone."""
    return datetime.datetime.now(CAMBODIA_TZ)

def _ensure_user_data(user: 'telegram.User'):
    """Creates a data entry for a user if it doesn't exist."""
    user_id = user.id
    if user_id not in user_data:
        user_data[user_id] = {
            "name": user.full_name,
            "check_in": None,
            "check_out": None,
            "wc_count": 0, "wc_late": 0,
            "smoke_count": 0, "smoke_late": 0,
            "eat_count": 0, "eat_late": 0
        }

async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Logs the error and sends a telegram message to notify the developer."""
    logger.error("Exception while handling an update:", exc_info=context.error)

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Sends a welcome message when the /start command is issued."""
    await update.message.reply_text(
        "Welcome to the Time Tracker Bot!\n\n"
        "You can use the following commands:\n"
        "- `check in`\n"
        "- `check out`\n"
        "- `wc`\n"
        "- `smoke`\n"
        "- `eat`\n\n"
        "When you are back from a break, please reply with `1`, `+1`, `back`, `finish`, or `done`."
    )

async def check_in(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'check in' message."""
    user = update.message.from_user
    _ensure_user_data(user) 

    now = get_now()
    user_data[user.id]["check_in"] = now

    if user.username:
        logger.info(f"User @{user.username} checking in. Attempting to open owner.")
        await _set_owner_status_in_db(user.username, is_stopped=False)
    else:
        logger.warning(f"User {user.full_name} (ID: {user.id}) checked in but has no Telegram @username. Cannot link to owner.")

    if datetime.time(13, 0) <= now.time() < datetime.time(15, 0):
        await update.message.reply_text("Well done!")
    elif datetime.time(15, 5) < now.time() < datetime.time(21, 0):
        late_minutes = int((now - now.replace(hour=15, minute=0, second=0, microsecond=0)).total_seconds() / 60)
        await update.message.reply_text(f"You are late {late_minutes} minutes.")
    elif datetime.time(21, 1) < now.time() < datetime.time(21, 59):
        late_minutes = int((now - now.replace(hour=21, minute=0, second=0, microsecond=0)).total_seconds() / 60)
        await update.message.reply_text(f"You are late {late_minutes} minutes (from 9 PM).")

async def check_out(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'check out' message."""
    user = update.message.from_user 
    user_id = user.id
    now = get_now()

    checkout_break_start = datetime.time(21, 0)
    checkout_break_end = datetime.time(23, 59, 59)
    checkout_final_start = datetime.time(3, 0)
    checkout_final_end = datetime.time(6, 0)

    is_valid_time = (checkout_break_start <= now.time() <= checkout_break_end) or \
                      (checkout_final_start <= now.time() <= checkout_final_end)

    if is_valid_time:
        if user_id in user_data:
            user_data[user_id]["check_out"] = now
        
        if user.username:
            logger.info(f"User @{user.username} checking out. Attempting to stop owner.")
            await _set_owner_status_in_db(user.username, is_stopped=True)
        else:
            logger.warning(f"User {user.full_name} (ID: {user.id}) checked out but has no Telegram @username. Cannot link to owner.")

    else:
        await update.message.reply_text("You are not allowed to check out at this time.")

async def wc(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'wc' message."""
    user = update.message.from_user
    _ensure_user_data(user) 

    if user.id not in user_breaks:
        user_breaks[user.id] = {"type": "wc", "start_time": get_now()}
    else:
        await update.message.reply_text("You are already on another break.")


async def smoke(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'smoke' message."""
    user = update.message.from_user
    _ensure_user_data(user)

    if user.id not in user_breaks:
        user_breaks[user.id] = {"type": "smoke", "start_time": get_now()}
    else:
        await update.message.reply_text("You are already on another break.")


async def eat(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles the 'eat' message."""
    user = update.message.from_user
    now = get_now()

    eat_time1_start = now.replace(hour=17, minute=0, second=0, microsecond=0)
    eat_time1_end = now.replace(hour=17, minute=30, second=0, microsecond=0)
    eat_time2_start = now.replace(hour=00, minute=30, second=0, microsecond=0)
    eat_time2_end = now.replace(hour=1, minute=00, second=0, microsecond=0)

    if (eat_time1_start <= now <= eat_time1_end) or \
       (eat_time2_start <= now <= eat_time2_end):

        _ensure_user_data(user) 

        if user.id not in user_breaks:
            user_breaks[user.id] = {"type": "eat", "start_time": get_now()}
        else:
            await update.message.reply_text("You are already on another break.")
    else:
        await update.message.reply_text("It's not time to eat yet.")

async def back_from_break(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Handles replies for returning from breaks and calculates lateness."""
    user = update.message.from_user
    user_id = user.id
    
    if user_id in user_breaks:
        break_info = user_breaks.pop(user_id)
        break_type = break_info["type"]
        start_time = break_info["start_time"]
        end_time = get_now()

        _ensure_user_data(user) 

        late_minutes = 0

        if break_type == 'wc':
            user_data[user_id]["wc_count"] += 1
            duration_minutes = (end_time - start_time).total_seconds() / 60
            if duration_minutes > 15:
                late_minutes = int(duration_minutes - 15)
                user_data[user_id]["wc_late"] += late_minutes
        
        elif break_type == 'smoke':
            user_data[user_id]["smoke_count"] += 1
            duration_minutes = (end_time - start_time).total_seconds() / 60
            if duration_minutes > 10:
                late_minutes = int(duration_minutes - 10)
                user_data[user_id]["smoke_late"] += late_minutes

        elif break_type == 'eat':
            user_data[user_id]["eat_count"] += 1
            
            deadline = None
            
            if start_time.hour == 17:
                deadline = start_time.replace(hour=17, minute=30, second=0, microsecond=0)
            
            elif start_time.hour == 1:
                deadline = start_time.replace(hour=00, minute=30, second=0, microsecond=0)

            if deadline and end_time > deadline:
                late_minutes = int((end_time - deadline).total_seconds() / 60)
                if late_minutes > 0:
                    user_data[user_id]["eat_late"] += late_minutes
        
        if late_minutes > 0:
            await update.message.reply_text(f"You are late {late_minutes} minutes.")

async def get_report(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Generates and sends the daily report if requested by an admin."""
    
    user = update.message.from_user
    if user.username != ADMIN_USERNAME:
        await update.message.reply_text("You are not authorized to perform this action.")
        return

    if not user_data:
        await update.message.reply_text("No activity to report for today.")
        return
        
    try:
        import pandas as pd
    except ImportError:
        logger.error("Pandas not installed. Cannot generate Excel report.")
        await update.message.reply_text("Report generation failed (missing library).")
        return
        
    sorted_users = sorted(user_data.items(), key=lambda item: item[1].get('check_in') or datetime.datetime.max.replace(tzinfo=CAMBODIA_TZ))

    report_data = []
    for user_id, data in sorted_users:
        check_in_time = data["check_in"].strftime("%H:%M") if data.get("check_in") else ""
        check_out_time = data["check_out"].strftime("%H:%M") if data.get("check_out") else ""
        
        report_data.append({
            "User": data["name"],
            "Check-in": check_in_time,
            "Check-out": check_out_time,
            "WC": data["wc_count"],
            "WC late (m)": data["wc_late"],
            "Smoke": data["smoke_count"],
            "Smoke late (m)": data["smoke_late"],
            "Eat": data["eat_count"],
            "Eat late (m)": data["eat_late"],
        })
    
    df = pd.DataFrame(report_data)
    file_path = f"daily_report_{get_now().strftime('%Y-%m-%d')}.xlsx"
    df.to_excel(file_path, index=False)
    
    
    wb = load_workbook(file_path)
    ws = wb.active

    header_fill = PatternFill(start_color="4AACC5", end_color="4AACC5", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    late_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    center_align = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = thin_border
        cell.alignment = center_align

    late_columns_indices = [col_idx for col_idx, cell in enumerate(ws[1], 1) if cell.value in ["WC late (m)", "Smoke late (m)", "Eat late (m)"]]

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = thin_border
            cell.alignment = center_align
            
            if cell.column in late_columns_indices and isinstance(cell.value, (int, float)) and cell.value > 0:
                cell.fill = late_fill
                cell.value = f"{cell.value} minute"

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = (max_length + 2)

    wb.save(file_path)
    

    with open(file_path, 'rb') as document:
        await context.bot.send_document(chat_id=update.message.chat_id, document=document)
    
    os.remove(file_path)
    await update.message.reply_text("Report sent.")


async def clear_data_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    """Clears all user data and STOPS all owners in the other bot."""
    
    await _stop_all_owners_in_db()
    
    user_data.clear()
    user_breaks.clear()
    logger.info("Daily user data has been cleared automatically.")

async def post_initialization(application: Application):
    """Runs once after the bot is initialized to setup DB."""
    await get_db_pool()

async def post_shutdown(application: Application):
    """Runs once before the bot shuts down to close DB."""
    await close_db_pool()

def main() -> None:
    """Start the bot."""
    if not DATABASE_URL:
        logger.critical("DATABASE_URL environment variable is not set. Bot cannot start.")
        return

    application = (
        Application.builder()
        .token(TELEGRAM_TOKEN)
        .post_init(post_initialization)  
        .post_shutdown(post_shutdown) 
        .build()
    )
    application.add_error_handler(error_handler)
    
    job_queue = application.job_queue
    clear_time = datetime.time(hour=4, minute=00, tzinfo=CAMBODIA_TZ)
    job_queue.run_daily(clear_data_job, clear_time)
    
    CHECKIN_REGEX = re.compile(r"^\s*(?:check\s*[- ]?in|checkin|ci|in|start(?:\s*[- ]?work)?)\s*$", re.IGNORECASE)
    CHECKOUT_REGEX = re.compile(r"^\s*(?:check\s*[- ]?out|checkout|co|out|end(?:\s*[- ]?work)?)\s*$", re.IGNORECASE)
    WC_REGEX = re.compile(r"^\s*(?:wc|toilet|restroom)(?:\d{1,2})?\s*$", re.IGNORECASE)
    SMOKE_REGEX = re.compile(r"^\s*(?:sm|smoke|cig(?:arette)?)(?:\d{1,2})?\s*$", re.IGNORECASE)
    EAT_REGEX = re.compile(r"^\s*(?:eat|meal|dinner|lunch)(?:\d{1,2})?\s*$", re.IGNORECASE)
    END_TOKENS_REGEX = re.compile(r"^\s*(?:\+?1|back(?:\s+to\s+seat)?|finish|done)\s*$", re.IGNORECASE)

    application.add_handler(CommandHandler("start", start))
    application.add_handler(CommandHandler("getreport", get_report))

    application.add_handler(MessageHandler(filters.Regex(CHECKIN_REGEX), check_in))
    application.add_handler(MessageHandler(filters.Regex(CHECKOUT_REGEX), check_out))
    application.add_handler(MessageHandler(filters.Regex(WC_REGEX), wc))
    application.add_handler(MessageHandler(filters.Regex(SMOKE_REGEX), smoke))
    application.add_handler(MessageHandler(filters.Regex(EAT_REGEX), eat))
    application.add_handler(MessageHandler(filters.Regex(END_TOKENS_REGEX), back_from_break))
    
    application.run_polling()

if __name__ == '__main__':
    main()
